"""
parse_war_room.py
─────────────────────────────────────────────────────
Универсальный парсер данных War Room из XLSX-файлов,
скачанных через Google Drive MCP (download_file_content).

ИСПОЛЬЗОВАНИЕ (в Claude):
  1. Вызвать download_file_content для Авито → получить путь к result-файлу
  2. Вызвать download_file_content для VK/Юла → получить путь к result-файлу
  3. Запустить:
       python3 parse_war_room.py <avito_result.txt> <vk_result.txt> [дата_начала] [дата_конца]
     Даты в формате DD.MM.YYYY (например: 29.06.2026 05.07.2026)

  Без дат — берёт последние 10 дней для VK/Юла и последний лист для Авито.
"""

import sys, json, base64, io
from datetime import datetime, timedelta
import openpyxl


# ─── АВИТО ────────────────────────────────────────────────────────────────────

def load_xlsx(result_path):
    with open(result_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    xlsx_bytes = base64.b64decode(data['content'])
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)


def find_latest_sheet(wb):
    """Найти вкладку с наибольшей (последней) датой начала периода ДД.ММ-ДД.ММ"""
    import re
    best_date = None
    best_name = None
    current_year = datetime.now().year
    for name in wb.sheetnames:
        m = re.match(r'^(\d{1,2})\.(\d{2})', name.strip())
        if m:
            d, mo = int(m.group(1)), int(m.group(2))
            dt = datetime(current_year, mo, d)
            if best_date is None or dt > best_date:
                best_date = dt
                best_name = name
    return best_name


def parse_avito(result_path):
    wb = load_xlsx(result_path)
    print(f"\n📋 Авито — вкладки: {wb.sheetnames}")

    sheet_name = find_latest_sheet(wb)
    if not sheet_name:
        print("  ❌ Вкладка с периодом не найдена")
        return None

    ws = wb[sheet_name]
    print(f"  ✅ Читаем последний период: '{sheet_name}'")

    rows = list(ws.iter_rows(values_only=True))

    # Найти строку с заголовками (содержит «план» и «факт»)
    header_row = None
    for i, row in enumerate(rows):
        joined = ' '.join(str(v or '') for v in row).lower()
        if 'план' in joined and 'факт' in joined:
            header_row = i
            break

    # Проекты War Room — собираем строки с данными
    projects = []
    war_room_names = [
        'инвентаризация', 'курьеры', 'лента крд', 'лента кзн', 'магнит крд',
        'магнит кзн', 'ростикс', 'лента члб', 'лукойл', 'пальмира', 'лента спб',
        'лента мск', 'командор', 'оби'
    ]

    for row in rows:
        name_cell = str(row[0] or '').strip()
        name_lower = name_cell.lower()
        if not any(kw in name_lower for kw in war_room_names):
            continue
        vals = [v for v in row]
        if not any(v not in (None, 0, '', 0.0) for v in vals[1:5]):
            continue

        def num(v):
            try:
                f = float(v)
                return round(f) if f == int(f) else round(f, 2)
            except (TypeError, ValueError):
                return None

        plan = num(vals[1]) if len(vals) > 1 else None
        fact = num(vals[2]) if len(vals) > 2 else None
        cpl_prev = num(vals[4]) if len(vals) > 4 else None
        cpl_curr = num(vals[5]) if len(vals) > 5 else None
        spend_prev = num(vals[7]) if len(vals) > 7 else None
        spend_curr = num(vals[8]) if len(vals) > 8 else None

        projects.append({
            'name': name_cell,
            'plan': plan,
            'fact': fact,
            'cpl_prev': cpl_prev,
            'cpl': cpl_curr,
            'spend_prev': spend_prev,
            'spend': spend_curr,
        })

    # Найти спецблоки (УРБАН ДИКСИ, ЧЛБ ОБ)
    special = {}
    for i, row in enumerate(rows):
        row_str = ' '.join(str(v or '') for v in row)
        for key in ['УРБАН ДИКСИ', 'ЧЛБ ОБ']:
            if key in row_str and key not in special:
                def num(v):
                    try:
                        f = float(str(v).strip().replace('%',''))
                        return round(f, 4) if isinstance(f, float) else f
                    except:
                        return None
                # Конверсия: в таблице это доля (0.1197) либо строка "46,33%".
                # В дашборде нужны проценты (11.97) — нормализуем.
                def conv_pct(v):
                    if v is None: return None
                    s = str(v).strip().replace('%', '').replace(',', '.')
                    try:
                        f = float(s)
                    except ValueError:
                        return None
                    # доля < 1 → в проценты; уже проценты (>1) — как есть
                    return round(f * 100, 2) if f <= 1 else round(f, 2)

                leads = num(row[1]) if len(row) > 1 else None
                cpl   = num(row[2]) if len(row) > 2 else None
                spend = num(row[3]) if len(row) > 3 else None
                adapt = num(row[4]) if len(row) > 4 else None
                conv  = conv_pct(row[5]) if len(row) > 5 else None
                # Узнаём период (из заголовка над блоком)
                period_label = ''
                if i > 0:
                    above = ' '.join(str(v or '') for v in rows[i-1])
                    if 'Лиды' in above or 'лиды' in above:
                        period_label = above.strip()
                special[key] = {
                    'leads': leads, 'cpl': cpl, 'spend': spend,
                    'adapted': adapt, 'conversion': conv,
                    'period': period_label
                }

    # Итого. План недели = сумма планов проектов (отдельной строки ИТОГО в таблице нет).
    total_plan   = sum(p['plan'] or 0 for p in projects)
    total_fact   = sum(p['fact'] or 0 for p in projects)
    total_spend  = sum(p['spend'] or 0 for p in projects)
    avg_cpl      = round(total_spend / total_fact) if total_fact else None
    plan_pct     = round(total_fact / total_plan * 100) if total_plan else None

    result = {
        'sheet': sheet_name,
        'projects': projects,
        'special': special,
        'total': {'plan': total_plan, 'fact': total_fact, 'spend': total_spend,
                  'cpl': avg_cpl, 'plan_pct': plan_pct}
    }

    print(f"\n  Проекты ({len(projects)}):")
    for p in projects:
        status = ''
        if p['plan'] and p['fact']:
            pct = round(p['fact'] / p['plan'] * 100)
            status = f" ({pct}% плана)"
        print(f"    {p['name']}: факт={p['fact']} план={p['plan']} CPL={p['cpl']}₽ расход={p['spend']}₽{status}")

    print(f"\n  ИТОГО: план {total_plan} | факт {total_fact} лидов ({plan_pct}% плана) | {total_spend}₽ | CPL {avg_cpl}₽")

    for key, s in special.items():
        print(f"\n  {key} [{s.get('period','')}]:")
        print(f"    лиды={s['leads']} CPL={s['cpl']}₽ расход={s['spend']}₽ адапт={s['adapted']} конверсия={s['conversion']}")

    return result


# ─── VK / ЮЛА ─────────────────────────────────────────────────────────────────

def parse_vk_yula(result_path, date_from=None, date_to=None):
    wb = load_xlsx(result_path)
    print(f"\n📋 VK/Юла — вкладки: {wb.sheetnames}")

    # Найти лист с лидами
    leads_sheet = None
    for name in wb.sheetnames:
        ws = wb[name]
        row1 = [str(c.value or '').lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        joined = ' '.join(row1)
        if ('источник' in joined or 'канал' in joined) and 'дата' in joined:
            leads_sheet = name
            break

    if not leads_sheet:
        print("  ❌ Лист с лидами не найден")
        return None

    ws = wb[leads_sheet]
    print(f"  ✅ Лист с лидами: '{leads_sheet}'")

    headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    src_col = date_col = -1
    for i, h in enumerate(headers):
        hl = h.lower()
        if 'источник' in hl or 'канал' in hl:
            src_col = i
        if 'дата' in hl:
            date_col = i

    if src_col == -1 or date_col == -1:
        print(f"  ❌ Колонки не найдены. Заголовки: {headers}")
        return None

    # Диапазон дат
    today = datetime.now()
    if date_from is None:
        date_from = today - timedelta(days=10)
    if date_to is None:
        date_to = today

    vk_count = yula_count = other_count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        date_val = row[date_col] if len(row) > date_col else None
        if not date_val:
            continue
        if isinstance(date_val, datetime):
            lead_date = date_val
        else:
            try:
                parts = str(date_val).split('.')
                if len(parts) >= 2:
                    d, m = int(parts[0]), int(parts[1])
                    y = int(parts[2]) if len(parts) > 2 else today.year
                    if y < 100: y += 2000
                    lead_date = datetime(y, m, d)
                else:
                    continue
            except:
                continue

        if not (date_from.date() <= lead_date.date() <= date_to.date()):
            continue

        src = str(row[src_col] or '').strip().lower() if len(row) > src_col else ''
        if 'вк' in src or 'vk' in src or 'вконтакте' in src:
            vk_count += 1
        elif 'юла' in src or 'yula' in src or 'авито' in src:
            yula_count += 1
        else:
            other_count += 1

    total = vk_count + yula_count + other_count
    period_str = f"{date_from.strftime('%d.%m')}–{date_to.strftime('%d.%m')}"

    result = {
        'period': period_str,
        'vk': vk_count,
        'yula': yula_count,
        'other': other_count,
        'total': total
    }

    print(f"\n  Период: {period_str}")
    print(f"  ВК: {vk_count} | Юла: {yula_count}" + (f" | Прочие: {other_count}" if other_count else ''))
    print(f"  ИТОГО: {total} лидов")

    return result


# ─── ЗАПУСК ───────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    args = sys.argv[1:]

    avito_path = args[0] if len(args) > 0 else None
    vk_path    = args[1] if len(args) > 1 else None

    date_from = date_to = None
    if len(args) > 3:
        try:
            date_from = datetime.strptime(args[2], '%d.%m.%Y')
            date_to   = datetime.strptime(args[3], '%d.%m.%Y')
        except ValueError:
            print(f"⚠️  Неверный формат дат. Используем последние 10 дней.")

    print("=" * 60)
    print("WAR ROOM DATA PARSER")
    print("=" * 60)

    if avito_path:
        parse_avito(avito_path)

    if vk_path:
        parse_vk_yula(vk_path, date_from, date_to)
